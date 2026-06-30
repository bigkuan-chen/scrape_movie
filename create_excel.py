import os
import sys
import io
import requests
import urllib3
from bs4 import BeautifulSoup
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Force stdout to UTF-8 to prevent encoding crashes on Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Suppress urllib3 SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def main():
    # 1. Parse HTML from page1.html
    html_file = "page1.html"
    if not os.path.exists(html_file):
        print(f"Error: {html_file} not found.")
        return
        
    with open(html_file, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f.read(), "html.parser")
        
    movie_cards = soup.select(".el-card.item")
    
    # Create directories for posters
    os.makedirs("posters", exist_ok=True)
    os.makedirs("posters/resized", exist_ok=True)
    
    movies = []
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    
    print("Downloading posters and parsing details...")
    for idx, card in enumerate(movie_cards, 1):
        # Title
        title_h2 = card.select_one("h2")
        title = title_h2.get_text(strip=True) if title_h2 else f"Movie_{idx}"
        
        # Split Title
        title_cn = title
        title_en = ""
        if " - " in title:
            parts = title.split(" - ", 1)
            title_cn = parts[0].strip()
            title_en = parts[1].strip()
            
        # Categories
        cat_spans = card.select(".categories button span")
        categories = ", ".join([span.get_text(strip=True) for span in cat_spans])
        
        # Info
        info_divs = card.select(".info")
        regions = ""
        duration = ""
        release_date = ""
        if len(info_divs) > 0:
            info_text = info_divs[0].get_text(strip=True)
            if "/" in info_text:
                parts = info_text.split("/", 1)
                regions = parts[0].strip()
                duration = parts[1].strip()
            else:
                regions = info_text
        if len(info_divs) > 1:
            release_date = info_divs[1].get_text(strip=True).replace(" 上映", "").strip()
            
        # Score
        score_p = card.select_one(".score")
        score = float(score_p.get_text(strip=True)) if score_p else 0.0
        
        # Cover Image URL
        cover_img = card.select_one("img.cover")
        cover_url = cover_img.get("src", "") if cover_img else ""
        if cover_url and "@" in cover_url:
            cover_url = cover_url.split("@")[0]
        
        # Detail Link
        detail_a = card.select_one("a[href*='/detail/']")
        detail_link = ""
        if detail_a:
            href = detail_a.get("href", "")
            detail_link = f"https://ssr1.scrape.center{href}" if href.startswith("/") else href
            
        # Download and resize poster
        poster_path = f"posters/movie_{idx}.jpg"
        resized_path = f"posters/resized/movie_{idx}.jpg"
        
        if cover_url:
            try:
                print(f"Downloading poster for {title_cn}...")
                resp = requests.get(cover_url, headers=headers, verify=False, timeout=15)
                if resp.status_code == 200:
                    with open(poster_path, "wb") as f_img:
                        f_img.write(resp.content)
                    
                    # Resize to fit Excel cell cleanly (width=45, height=60)
                    with PILImage.open(poster_path) as im:
                        im = im.convert("RGB")  # Ensure it's RGB mode
                        im.thumbnail((45, 60))
                        im.save(resized_path, "JPEG")
                else:
                    resized_path = None
            except Exception as e:
                print(f"Failed to download/resize poster for {title_cn}: {e}")
                resized_path = None
        else:
            resized_path = None
            
        movies.append({
            "title_cn": title_cn,
            "title_en": title_en,
            "categories": categories,
            "regions": regions,
            "duration": duration,
            "release_date": release_date,
            "score": score,
            "detail_link": detail_link,
            "poster_path": resized_path
        })
        
    # 2. Create Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Movies"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Headers
    headers_list = [
        "Poster", "Chinese Title", "English/Original Title", "Categories",
        "Regions/Countries", "Duration", "Release Date", "Score", "Detail Link"
    ]
    
    # Write and style headers
    for col_idx, h in enumerate(headers_list, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = h
        cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")  # Steel blue fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ws.row_dimensions[1].height = 28
    
    # Set custom column dimensions to accommodate poster images and titles
    ws.column_dimensions['A'].width = 8   # Poster
    ws.column_dimensions['B'].width = 22  # Chinese Title
    ws.column_dimensions['C'].width = 25  # English Title
    ws.column_dimensions['D'].width = 22  # Categories
    ws.column_dimensions['E'].width = 22  # Regions/Countries
    ws.column_dimensions['F'].width = 12  # Duration
    ws.column_dimensions['G'].width = 14  # Release Date
    ws.column_dimensions['H'].width = 8   # Score
    ws.column_dimensions['I'].width = 35  # Detail Link
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Fill row data
    for row_idx, m in enumerate(movies, 2):
        ws.row_dimensions[row_idx].height = 50  # Fit 60px height image
        
        # Insert Poster Image into Column A
        if m["poster_path"] and os.path.exists(m["poster_path"]):
            img = OpenpyxlImage(m["poster_path"])
            ws.add_image(img, f"A{row_idx}")
            
        # Write text values
        ws.cell(row=row_idx, column=2, value=m["title_cn"])
        ws.cell(row=row_idx, column=3, value=m["title_en"])
        ws.cell(row=row_idx, column=4, value=m["categories"])
        ws.cell(row=row_idx, column=5, value=m["regions"])
        ws.cell(row=row_idx, column=6, value=m["duration"])
        ws.cell(row=row_idx, column=7, value=m["release_date"])
        ws.cell(row=row_idx, column=8, value=m["score"])
        
        # Add hyperlink to Detail Link
        cell_link = ws.cell(row=row_idx, column=9)
        cell_link.value = m["detail_link"]
        if m["detail_link"]:
            cell_link.hyperlink = m["detail_link"]
            cell_link.font = Font(name="Microsoft YaHei", size=10, color="0563C1", underline="single")
            
        # Standardize alignments and borders
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx > 1:
                cell.font = Font(name="Microsoft YaHei", size=10)
                if col_idx in [6, 7, 8]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx == 9:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
    # Save the file
    excel_file = "movies_with_posters.xlsx"
    wb.save(excel_file)
    print(f"\nExcel file successfully generated and saved to: {excel_file}")

if __name__ == "__main__":
    main()
